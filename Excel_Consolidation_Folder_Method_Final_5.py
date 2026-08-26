import os
from datetime import datetime
import shutil
import json
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage

from PIL import Image, ImageChops
import math
import pytesseract
from bs4 import BeautifulSoup
import imagehash

pytesseract.pytesseract.tesseract_cmd = (
    r"C:\Program Files\Tesseract-OCR\tesseract.exe"
)


# =====================================================
# CONFIGURATION
# =====================================================


ROOT_FOLDER = r"C:\Python\HRC"
OUTPUT_FILE = ROOT_FOLDER + r"\Combined.xlsx"
HASH_THRESHOLD = 4

SUPPORTED_IMAGES = (
    ".png",
    ".jpg",
    ".jpeg",
    ".bmp",
    ".gif"
)

reference_hashes = {}
all_hashes = {}
filtered_hashes = {}


def image_difference(img1_path, img2_path):
  img1 = Image.open(img1_path).convert("L").resize((300, 200))
  img2 = Image.open(img2_path).convert("L").resize((300, 200))

  diff = ImageChops.difference(img1, img2)

  histogram = diff.histogram()

  rms = math.sqrt(
  sum(value * ((idx % 256) ** 2)
  for idx, value in enumerate(histogram)) 
/ float(img1.size[0] * img1.size[1])
)

  return rms


# =====================================================
# HELPERS
# =====================================================

def get_unique_sheet_name(workbook, base_name):

    base_name = str(base_name).replace("/", "_")
    base_name = str(base_name).replace("\\", "_")
    base_name = base_name[:31]

    if base_name not in workbook.sheetnames:
        return base_name

    counter = 1

    while True:

        candidate = f"{base_name[:25]}_{counter}"

        if candidate not in workbook.sheetnames:
            return candidate

        counter += 1


def find_folder(root_path, target_folder):

    for current_root, dirs, files in os.walk(root_path):

        for d in dirs:

            if d.lower() == target_folder.lower():

                return os.path.join(current_root, d)

    return None


def read_excel_file(file_path):

    try:

        if file_path.lower().endswith(".xlsx"):
            return pd.read_excel(
                file_path,
                engine="openpyxl"
            )

        try:
            return pd.read_csv(
                file_path,
                sep="\t",
                engine="python"
            )

        except:
            return pd.read_excel(
                file_path,
                engine="xlrd"
            )

    except Exception as e:

        print(f"\nFailed reading: {file_path}")
        print(e)

        return None




def create_business_screenshots_hash(folder_path, screenshots_path):

    images = []

    for file in os.listdir(screenshots_path):

        if file.lower().endswith(SUPPORTED_IMAGES):

            full_path = os.path.join(
                screenshots_path,
                file
            )

            images.append(
                (
                    full_path,
                    file,
                    os.path.getmtime(full_path)
                )
            )

    images.sort(key=lambda x: x[2])

    # -----------------------------------
    # Remove duplicates
    # -----------------------------------

    filtered_images = []
    if len(filtered_hashes)>0:
        filtered_hashes.clear()
    for img in images:
        matched, img_hash = find_matching_image_in_all(img[0])
        if matched:
            matched1, matching_file1 = find_matching_image_in_filtered(img[0])
            if matched1:
                continue
            else:
                filtered_images.append(img)
                filtered_hashes[img[0]] = img_hash
        else:
            filtered_images.append(img)
            filtered_hashes[img[0]] = img_hash
    print(f"Loaded {len(filtered_hashes)} filtered images")    

    # -----------------------------------
    # Skip login screens
    # -----------------------------------

    business_images = []

    for img_info in filtered_images:

       # try:

        img = Image.open(img_info[0])

        text = pytesseract.image_to_string(
            img
        ).lower()
        
        filename = Path(img_info[0]).name
        if filename == "screenshot_091.png" or filename == "screenshot_095.png" or filename == "screenshot_097.png" or filename == "screenshot_100.png": 
            print(text)
        skip_words = [
            "oracle e-business suite",
            "user name",
            "password",
            "navigator",
            "oracle forms",
            "do you want to run this application",
            "duckduckgo",
            "search privately",
            "new tab",
            "search the web",
            "sign in",
            "log in",
            "login",
            "browser",
            "address bar",
            "favorites",
            "bookmarks",
            "order to cash",
            "loading"

        ]

        skip_screen = any(
            word in text
            for word in skip_words
        )

        if not skip_screen:
            business_images.append(img_info)
            reference_hashes[img_info[0]] = imagehash.phash(img)

       # except Exception as e:

        # print(f"OCR failed for {img_info[1]}")
        # print(e)     

# =====================================================
# SCREENSHOTS SHEET
# =====================================================

def add_screenshot_sheet( workbook, folder_name):
    
    parent_folder = Path(folder_name).parent
    sheet_prefix = get_business_sheet_name(parent_folder)
    if sheet_prefix == None:
        sheet_prefix = parent_folder.name
    sheet_name = get_unique_sheet_name(
        workbook,
        f"{sheet_prefix}_Screenshots"
    )

    ws = workbook.create_sheet(
        title=sheet_name
    )
    HTML_FILE = str(parent_folder)+r"\test-report.html"
    with open(Path(HTML_FILE), "r", encoding="utf-8") as f:
        soup = BeautifulSoup(f, "html.parser")

    # ws["A1"] = "Step Action"
    # ws["B1"] = "Matched Reference Image"

    current_row = 2

    # =========================
    # Process step-cards
    # =========================

    step_cards = soup.find_all(class_="step-card")

    for card in step_cards:
        step_div = card.find("span", class_="step-name" )
        step_text = step_div.get_text(" ", strip=True) if step_div else ""
        action_div = card.find("div", class_="step-action")
        action_text = action_div.get_text(" ", strip=True) if action_div else ""

        visual_divs = card.find_all("div", class_="turn-visual")

        if not visual_divs:
            continue

        for visual_div in visual_divs:
            img_tag = visual_div.find("img")

            if not img_tag or not img_tag.get("src"):
                continue

            img_path = img_tag["src"]

            if not os.path.isabs(img_path):
                img_path = os.path.join(
                    os.path.dirname(os.path.abspath(HTML_FILE)),
                    img_path
                )

            if not os.path.exists(img_path):
                print(f"Image not found: {img_path}")
                continue

            # ==========================================
            # Compare image against reference folder
            # ==========================================
            matched, matching_file = find_matching_image(img_path)

            if not matched:
                print(f"No match found for {img_path}")
                continue

            # print(f"Matched: {img_path} --> {matching_file}")

            # ==========================================
            # Save to Excel ONLY if matched
            # ==========================================
            ws.cell(row=current_row, column=1, value=step_text)
            ws.cell(row=current_row, column=2, value=action_text)
            # ws.cell(row=current_row, column=2, value=os.path.basename(matching_file))

            try:
                excel_img = ExcelImage(img_path)

                max_width = 800

                if excel_img.width > max_width:
                    ratio = max_width / excel_img.width
                    excel_img.width = int(excel_img.width * ratio)
                    excel_img.height = int(excel_img.height * ratio)

                image_row = current_row + 1

                ws.add_image(excel_img, f"A{image_row}")

                # ws.row_dimensions[image_row].height = 150

                current_row += 35

            except Exception as e:
                print(f"Excel image insert failed: {e}")
                current_row += 2


    # =========================
    # Save Workbook
    # =========================

    # ws.column_dimensions["A"].width = 80
    # ws.column_dimensions["B"].width = 40

    # workbook.save()

    print(f"Excel sheet added: {sheet_name}")
        
    

def add_screenshot_sheet1(
        workbook,
        folder_name,
        screenshots_path):

    images = []

    for file in os.listdir(screenshots_path):

        if file.lower().endswith(SUPPORTED_IMAGES):

            full_path = os.path.join(
                screenshots_path,
                file
            )

            images.append(
                (
                    full_path,
                    file,
                    os.path.getmtime(full_path)
                )
            )

    if not images:

        print(
            f"No screenshots found in {folder_name}"
        )

        return

    images.sort(key=lambda x: x[2])
    parent_folder = Path(screenshots_path).parent
    sheet_prefix = get_business_sheet_name(parent_folder)
    if sheet_prefix == None:
        sheet_prefix = folder_name
    sheet_name = get_unique_sheet_name(
        workbook,
        f"{sheet_prefix}_Screenshots"
    )

    ws = workbook.create_sheet(
        title=sheet_name
    )

    row = 1

    for img_path, filename, timestamp in images:

        ws.cell(
            row=row,
            column=1,
            # value=filename
        )

        excel_img = ExcelImage(img_path)

        excel_img.width = 800
        excel_img.height = 467

        ws.add_image(
            excel_img,
            f"A{row + 1}"
        )

        row += 26

    print(
        f"Added Screenshot Sheet: {sheet_name}"
    )

   
    
# =====================================================
# DOWNLOAD REPORT SHEETS
# =====================================================

def add_download_sheets(
        workbook,
        downloads_path,folder_name):

    for file in os.listdir(downloads_path):

        if not file.lower().endswith(
                (".xls", ".xlsx")):
            continue

        file_path = os.path.join(
            downloads_path,
            file
        )

        print(
            f"Reading report: {file}"
        )

        df = read_excel_file(file_path)

        if df is None:
            continue
        
        parent_folder = Path(downloads_path).parent
        sheet_prefix = get_business_sheet_name(parent_folder)
        if sheet_prefix == None:
            sheet_prefix = folder_name
        sheet_type = get_sheet_type(parent_folder)
        report_name = "default"
        if ("Number" in df.columns) and ("Balance" in df.columns):
            report_name = "Transactions"

        if ("Num" in df.columns) and ("Item" in df.columns) and ("Description" in df.columns) and ("UOM" in df.columns):
            report_name = "Line_Items"

        if ("Num" in df.columns) and ("Item" in df.columns) and ("Description" in df.columns) and ("UOM" in df.columns) and pd.to_numeric(df.stack(), errors='coerce').lt(0).any():
                report_name = "CMLine_Items"

        if report_name == "default": 
            report_name = os.path.splitext(file)[0] 


        x=False
        x = ((sheet_type == "WO") and ("export" in report_name)) 
        x = x or ((sheet_type == "SO") and ("Transactions" in report_name)) 
        x = x or ((sheet_type == "SO") and ("Transactions" in report_name)) 
        x = x or ((sheet_type == "Inv") and ("Line_Items" in report_name)) 
        x = x or ((sheet_type == "CM") and ("CMLine_Items" in report_name))
        x = x or ((sheet_type == "CCWR") and ("Terminate" in report_name)) 
        x = x or ((sheet_type == "Contract") and ("Contract" in report_name))

        if x:    
            sheet_name = get_unique_sheet_name(
                    workbook,
                    f"{sheet_prefix}_{report_name}"
                )
            ws = workbook.create_sheet(
                title=sheet_name
            )

            # Header

            for col_num, column_name in enumerate(
                    df.columns,
                    start=1):

                ws.cell(
                    row=1,
                    column=col_num,
                    value=str(column_name)
                )

            # Data

            for row_num, row_data in enumerate(
                    df.values.tolist(),
                    start=2):

                for col_num, value in enumerate(
                        row_data,
                        start=1):

                    ws.cell(
                        row=row_num,
                        column=col_num,
                        value=""
                        if pd.isna(value)
                        else str(value)
                    )

            print(
                f"Added report sheet: {sheet_name}"
            )

# =====================================================
# SHEETNAME MODIFY
# =====================================================


def get_business_sheet_name(folder_path):
    """
    Reads output-data.json from the specified folder and returns
    the business sheet name based on the JSON content.

    Rules:
    - If OrderNumber exists -> return "SO#<OrderNumber>"
    - Else if TransactionNumber exists -> return "<Type>#<TransactionNumber>"
    """

    json_file = os.path.join(folder_path, "output-data.json")

    with open(json_file, "r", encoding="utf-8") as f:
        data = json.load(f)

    if "orderNumber" in data and data["orderNumber"]:
        return f"CCW#{data['orderNumber']}"

    if "SalesOrder" in data and data["SalesOrder"]:
        return f"SO#{data['SalesOrder']}"

    if "TransactionNumber" in data and data["TransactionNumber"]:
        return f"{data.get('invoiceType', '')}#{data['TransactionNumber']}"

    if "CMTransactionNumber" in data and data["CMTransactionNumber"]:
        return f"{data.get('invoiceType', '')}#{data['CMTransactionNumber']}"

    if "CCWR_Transaction" in data and data["CCWR_Transaction"]:
        return f"CCWR#{data['CCWR_Transaction']}"

    if "ContractNumber" in data and data["ContractNumber"]:
        return f"Contract#{data['ContractNumber']}"
    

    return None

def get_sheet_type(folder_path):
    """
    Reads output-data.json from the specified folder and returns
    the business sheet name based on the JSON content.

    Rules:
    - If OrderNumber exists -> return "SO#<OrderNumber>"
    - Else if TransactionNumber exists -> return "<Type>#<TransactionNumber>"
    """

    json_file = os.path.join(folder_path, "output-data.json")

    with open(json_file, "r", encoding="utf-8") as f:
        data = json.load(f)

    if "orderNumber" in data and data["orderNumber"]:
        return "WO"

    if "SalesOrder" in data and data["SalesOrder"]:
        return "SO"

    if "TransactionNumber" in data and data["TransactionNumber"]:
        return "Inv"

    if "CMTransactionNumber" in data and data["CMTransactionNumber"]:
        return "CM"

    if "CCWR_Transaction" in data and data["CCWR_Transaction"]:
        return "CCWR"

    if "ContractNumber" in data and data["ContractNumber"]:
        return "Contract"
    

    return None



def find_matching_image_in_all(image_path):
    """
    Compare image against all reference images and return:
    (True, matching_file) if match found
    (False, None) otherwise
    """

    try:
        img_hash = imagehash.phash(Image.open(image_path))

        best_match = None
        best_distance = float("inf")

        for ref_file, ref_hash in all_hashes.items():
            distance = img_hash - ref_hash

            if distance < best_distance:
                best_distance = distance
                best_match = ref_file

        if best_distance <= HASH_THRESHOLD:
            return True, img_hash

        return False, img_hash

    except Exception as e:
        print(f"Comparison failed for {image_path}: {e}")
        return False, None

def find_matching_image_in_filtered(image_path):
    """
    Compare image against all reference images and return:
    (True, matching_file) if match found
    (False, None) otherwise
    """

    try:
        img_hash = imagehash.phash(Image.open(image_path))

        best_match = None
        best_distance = float("inf")

        for ref_file, ref_hash in filtered_hashes.items():
            distance = img_hash - ref_hash

            if distance < best_distance:
                best_distance = distance
                best_match = ref_file

        if best_distance <= HASH_THRESHOLD:
            return True, best_match

        return False, None

    except Exception as e:
        print(f"Comparison failed for {image_path}: {e}")
        return False, None


def find_matching_image(image_path):
    """
    Compare image against all reference images and return:
    (True, matching_file) if match found
    (False, None) otherwise
    """

    try:
        img_hash = imagehash.phash(Image.open(image_path))

        best_match = None
        best_distance = float("inf")

        for ref_file, ref_hash in reference_hashes.items():
            distance = img_hash - ref_hash

            if distance < best_distance:
                best_distance = distance
                best_match = ref_file

        if best_distance <= HASH_THRESHOLD:
            return True, best_match

        return False, None

    except Exception as e:
        print(f"Comparison failed for {image_path}: {e}")
        return False, None

# =====================================================
# MAIN
# =====================================================

def main():

    wb = Workbook()

    wb.remove(
        wb.active
    )

    print("\nScanning folders...\n")

    for folder in os.listdir(ROOT_FOLDER):

        folder_path = os.path.join(
            ROOT_FOLDER,
            folder
        )

        if not os.path.isdir(folder_path):
            continue

        print(
            "\n================================"
        )

        print(
            f"Processing Folder: {folder}"
        )

        screenshots_path = find_folder(
            folder_path,
            "screenshots"
        )

        downloads_path = find_folder(
            folder_path,
            "downloads"
        )

        print(
            f"Screenshots Path: {screenshots_path}"
        )

        print(
            f"Downloads Path: {downloads_path}"
        )

        # ----------------------------------
        # Screenshots
        # ----------------------------------

        if screenshots_path:
            if len(all_hashes) > 0:
                all_hashes.clear()
            for file in Path(screenshots_path).iterdir():
                if file.suffix.lower() in [".png", ".jpg", ".jpeg", ".bmp", ".gif", ".webp"]:
                    try:
                        img = Image.open(file)
                        all_hashes[str(file)] = imagehash.phash(img)
                    except Exception as e:
                        print(f"Error reading {file}: {e}")
    
            print(f"Loaded {len(all_hashes)} all images")

            if len(reference_hashes) > 0:
                reference_hashes.clear()
            create_business_screenshots_hash(
                folder_path,
                screenshots_path
                )
            
            print(f"Loaded {len(reference_hashes)} reference images")  

            add_screenshot_sheet(
                wb,
                screenshots_path,
                )

        
        # ----------------------------------
        # Excel Reports
        # ----------------------------------

        if downloads_path:

            add_download_sheets(
                wb,
                downloads_path,
                folder
            )

        else:

            print(
                "No downloads folder found"
            )

    wb.save(
        OUTPUT_FILE
    )

    print(
        "\nSUCCESS!"
    )

    print(
        f"Workbook created: {OUTPUT_FILE}"
    )


if __name__ == "__main__":
    main()